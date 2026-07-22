import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_excel_report(results_list, output_file="appium_automation/reports/Appium_Execution_Report.xlsx"):
    """
    Generates a professionally styled Excel report matching the MedMonitor AI layout.
    """
    # Ensure directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Convert list of dictionaries to DataFrame
    df = pd.DataFrame(results_list)
    
    # Check if df is empty
    if df.empty:
        print("No test results to report.")
        return

    # Write to Excel using pandas initially
    df.to_excel(output_file, index=False, sheet_name="Mobile Execution Results")
    
    # Load workbook with openpyxl to apply styling
    wb = load_workbook(output_file)
    ws = wb["Mobile Execution Results"]
    
    # Define MedMonitor AI Color Palette
    primary_color = "002C3E50" # Dark Blue/Slate
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True, name="Calibri", size=11)
    
    pass_fill = PatternFill(start_color="0027AE60", end_color="0027AE60", fill_type="solid") # Vibrant Green
    fail_fill = PatternFill(start_color="00E74C3C", end_color="00E74C3C", fill_type="solid") # Red
    skip_fill = PatternFill(start_color="00F39C12", end_color="00F39C12", fill_type="solid") # Orange/Yellow
    status_font = Font(color="FFFFFFFF", bold=True)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply Header Styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border

    # Get column index for "Status"
    status_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Status":
            status_col_idx = idx
            break

    # Apply Row Styling and Conditional Formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            
            # Highlight Status cell
            if status_col_idx and cell.column == status_col_idx:
                cell.alignment = alignment
                if cell.value == "Passed":
                    cell.fill = pass_fill
                    cell.font = status_font
                elif cell.value == "Failed":
                    cell.fill = fail_fill
                    cell.font = status_font
                elif cell.value == "Skipped":
                    cell.fill = skip_fill
                    cell.font = status_font

    # Adjust Column Widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 40:
            adjusted_width = 40 # Max width
        ws.column_dimensions[column].width = adjusted_width

    # Add Summary Section at the top
    ws.insert_rows(1, 4)
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "AuraFit AI - Appium Mobile Test Execution Report"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=primary_color)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    total_tests = len(results_list)
    passed_tests = len([r for r in results_list if r.get('Status') == 'Passed'])
    
    ws['A2'] = "Total Executed:"
    ws['B2'] = total_tests
    ws['A3'] = "Overall Status:"
    ws['B3'] = "PASS" if passed_tests == total_tests else "FAIL"
    ws['B3'].font = status_font
    ws['B3'].fill = pass_fill if passed_tests == total_tests else fail_fill
    ws['B3'].alignment = alignment

    wb.save(output_file)
    print(f"Excel report generated successfully at {os.path.abspath(output_file)}")

if __name__ == "__main__":
    # Test execution
    mock_data = [
        {"Test Case ID": "APP-TC-001", "Module": "Authentication", "Feature": "test_tap_login", "Device": "Pixel 6 Emulator", "Status": "Passed", "Duration (s)": 3.4},
    ]
    generate_excel_report(mock_data)
