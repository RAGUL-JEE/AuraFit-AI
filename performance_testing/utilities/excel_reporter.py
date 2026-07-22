import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def generate_excel_report(perf_summary, api_stats, vuln_stats, output_file="performance_testing/reports/Performance_Report.xlsx"):
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    wb = Workbook()
    
    # --- Sheet 1: Performance Summary ---
    ws1 = wb.active
    ws1.title = "Performance Summary"
    df_perf = pd.DataFrame(perf_summary)
    for r in dataframe_to_rows(df_perf, index=False, header=True):
        ws1.append(r)

    # --- Sheet 2: API Statistics ---
    ws2 = wb.create_sheet("API Statistics")
    df_api = pd.DataFrame(api_stats)
    for r in dataframe_to_rows(df_api, index=False, header=True):
        ws2.append(r)

    # --- Sheet 3: Vulnerability Assessment ---
    ws3 = wb.create_sheet("Vulnerability Assessment")
    df_vuln = pd.DataFrame(vuln_stats)
    for r in dataframe_to_rows(df_vuln, index=False, header=True):
        ws3.append(r)

    # --- Sheet 4: Graphs ---
    ws4 = wb.create_sheet("Graphs")
    
    # Bar Chart: Requests Per Second
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Requests Per Second (By Test)"
    chart1.y_axis.title = 'RPS'
    chart1.x_axis.title = 'Test Name'
    data1 = Reference(ws1, min_col=5, min_row=1, max_row=len(perf_summary)+1)
    cats1 = Reference(ws1, min_col=1, min_row=2, max_row=len(perf_summary)+1)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws4.add_chart(chart1, "A2")

    # Line Chart: Latency
    chart2 = LineChart()
    chart2.title = "Latency Trends (Avg, Min, Max)"
    chart2.style = 13
    chart2.y_axis.title = 'Latency (ms)'
    chart2.x_axis.title = 'Test Name'
    data2 = Reference(ws1, min_col=6, max_col=8, min_row=1, max_row=len(perf_summary)+1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats1)
    ws4.add_chart(chart2, "I2")

    # --- Formatting ---
    primary_color = "002C3E50" 
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True, name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws in [ws1, ws2, ws3]:
        # Format Headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border
            
        # Format Data Rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = alignment
                cell.border = thin_border
                
                # Conditional Formatting
                val = str(cell.value).lower()
                if "passed" in val or "not detected" in val:
                    cell.fill = PatternFill(start_color="0027AE60", end_color="0027AE60", fill_type="solid")
                    cell.font = Font(color="FFFFFFFF", bold=True)
                elif "failed" in val or "critical" in val or "high" in val:
                    cell.fill = PatternFill(start_color="00E74C3C", end_color="00E74C3C", fill_type="solid")
                    cell.font = Font(color="FFFFFFFF", bold=True)
                elif "warning" in val or "medium" in val:
                    cell.fill = PatternFill(start_color="00F39C12", end_color="00F39C12", fill_type="solid")
                    cell.font = Font(color="FFFFFFFF", bold=True)
                elif "low" in val or "informational" in val:
                    cell.fill = PatternFill(start_color="003498DB", end_color="003498DB", fill_type="solid")
                    cell.font = Font(color="FFFFFFFF", bold=True)

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)

    wb.save(output_file)
    print(f"Enterprise Excel generated successfully at {os.path.abspath(output_file)}")
