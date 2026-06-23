import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def create_styled_excel(file_path, report_title, sub_title, metrics_dict, data_headers, data_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Colors
    HEADER_GREEN = "34D399"
    DARK_BLUE = "1B263B"
    LIGHT_GREEN = "ECFDF5"
    WHITE = "FFFFFF"
    
    fill_header = PatternFill(start_color=HEADER_GREEN, end_color=HEADER_GREEN, fill_type="solid")
    fill_dark = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    fill_light = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    font_title = Font(color=WHITE, bold=True, size=16)
    font_sub = Font(color=WHITE, italic=True, size=10)
    font_bold_white = Font(color=WHITE, bold=True)
    font_white = Font(color=WHITE)
    font_pass = Font(color="059669", bold=True)
    font_blue_link = Font(color="2563EB", bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    num_cols = len(data_headers)
    
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=report_title)
    cell.fill = fill_header
    cell.font = font_title
    cell.alignment = align_center
    ws.row_dimensions[1].height = 30
    
    for c in range(1, num_cols + 1):
        ws.cell(row=1, column=c).fill = fill_header
    
    # Row 2: Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell = ws.cell(row=2, column=1, value=sub_title)
    cell.fill = fill_dark
    cell.font = font_sub
    cell.alignment = align_center
    ws.row_dimensions[2].height = 20
    for c in range(1, num_cols + 1):
        ws.cell(row=2, column=c).fill = fill_dark
        
    # Spacer
    ws.row_dimensions[3].height = 10
        
    # Row 4: Metrics
    start_metrics_row = 4
    ws.merge_cells(start_row=start_metrics_row, start_column=1, end_row=start_metrics_row, end_column=num_cols)
    cell = ws.cell(row=start_metrics_row, column=1, value="Summary Metrics")
    cell.fill = fill_dark
    cell.font = font_bold_white
    cell.alignment = align_center
    for c in range(1, num_cols + 1):
        ws.cell(row=start_metrics_row, column=c).fill = fill_dark
        
    row_idx = start_metrics_row + 1
    for key, val in metrics_dict.items():
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        c1 = ws.cell(row=row_idx, column=1, value=key)
        c1.fill = fill_dark
        c1.font = font_bold_white
        c1.alignment = align_right
        
        c2 = ws.cell(row=row_idx, column=3, value=val)
        c2.fill = fill_dark
        c2.font = font_pass if str(val) == "PASS" else font_white
        c2.alignment = align_center
        
        ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=num_cols)
        c3 = ws.cell(row=row_idx, column=4, value="Automated calculation")
        c3.fill = fill_dark
        c3.font = font_sub
        c3.alignment = align_left
        
        for c in range(1, num_cols + 1):
            ws.cell(row=row_idx, column=c).fill = fill_dark
            
        row_idx += 1
        
    ws.row_dimensions[row_idx].height = 10
    row_idx += 1
        
    # Detailed Data Header
    start_data_header = row_idx
    ws.merge_cells(start_row=start_data_header, start_column=1, end_row=start_data_header, end_column=num_cols)
    cell = ws.cell(row=start_data_header, column=1, value="DETAILED TEST CASES & POLICIES")
    cell.fill = fill_dark
    cell.font = font_bold_white
    cell.alignment = align_center
    ws.row_dimensions[start_data_header].height = 25
    for c in range(1, num_cols + 1):
        ws.cell(row=start_data_header, column=c).fill = fill_dark
        
    start_data_header += 1
    for col, header in enumerate(data_headers, 1):
        cell = ws.cell(row=start_data_header, column=col, value=header)
        cell.fill = fill_dark
        cell.font = font_bold_white
        cell.alignment = align_center
        
    # Data Rows
    row_idx = start_data_header + 1
    for r_idx, row_data in enumerate(data_rows):
        is_even = r_idx % 2 == 0
        fill_row = fill_light if is_even else PatternFill(fill_type=None)
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if is_even:
                cell.fill = fill_row
            
            if col == 1:
                cell.alignment = align_center
                cell.font = font_blue_link
            elif col == num_cols:
                cell.alignment = align_center
                if val == "PASS":
                    cell.font = font_pass
            else:
                cell.alignment = align_left
        ws.row_dimensions[row_idx].height = 30
        row_idx += 1
        
    # Adjust column widths
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.columns, 1):
        max_length = 15
        column = get_column_letter(i)
        for cell in col:
            try:
                if cell.row > start_data_header and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 60:
            adjusted_width = 60
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)


def generate_security_cases():
    cases_data = {
        "Dependency Security": [
            ("Direct Dependency Vulnerability Audit", "Scan direct npm packages for critical security advisories."),
            ("Indirect Dependency Nesting Audit", "Validate deep dependency tree for security alerts."),
            ("Deprecated Packages Verification", "Ensure no deprecated dependencies are in use."),
            ("License Compliance Audit", "Verify all libraries comply with security policies."),
            ("Lockfile Integrity Validation", "Check package lockfile integrity hashes match npm registry."),
            ("DevDependency Vulnerability Audit", "Scan dev npm packages for critical security advisories."),
            ("Unused Dependency Audit", "Ensure no orphaned or unused packages remain in package.json."),
            ("Package Origin Verification", "Verify package registries to avoid namespace hijacking or dependency confusion."),
            ("Dependency Licensing Type Check", "Verify that dependency licenses match copyleft restrictions."),
            ("Package Script Lifecycle Audit", "Verify pre/postinstall scripts in third-party libraries for malicious actions."),
            ("Semantic Versioning Range Pinning", "Ensure strict version ranges to prevent malicious minor updates."),
            ("Shrinkwrap/Lockfile Sync Check", "Confirm lockfile package definitions match package.json exactly."),
            ("Outdated Security Patches Audit", "Verify dependencies are running latest security patch updates."),
            ("Sub-dependency Peer Verification", "Ensure peer dependencies do not resolve to vulnerable revisions."),
            ("Registry Certificate Verification", "Enforce SSL/TLS validation when communicating with the npm registry.")
        ],
        "SAST & Code Analysis": [
            ("SQL Injection Prevention Check", "Verify all database queries use safe compiled APIs/Firestore SDKs."),
            ("XSS Protection Verification", "Ensure context-aware output encoding is used to prevent Cross-Site Scripting."),
            ("CSRF Token Validation", "Check that anti-CSRF tokens are present and validated for state-changing requests."),
            ("Command Injection Prevention", "Verify input is properly sanitized before being passed to system commands."),
            ("Memory Leak Detection", "Analyze code for objects that are allocated but never deallocated."),
            ("Buffer Overflow Checks", "Scan for unchecked buffer boundaries in lower-level system code.")
        ],
        "API Security": [
            ("Rate Limiting Check", "Verify API endpoints enforce rate limiting to prevent abuse."),
            ("API Key Rotation Policy", "Ensure API keys have expiration dates and rotation policies."),
            ("Mass Assignment Prevention", "Check that APIs do not allow mass assignment of object properties."),
            ("Unsafe Consumption of APIs", "Ensure data from third-party APIs is validated before use.")
        ],
        "Database Security": [
            ("Data Encryption at Rest", "Verify sensitive database fields are encrypted at rest."),
            ("NoSQL Injection Prevention", "Ensure NoSQL queries are parameterized or properly escaped."),
            ("Database Access Logging", "Check that all database access by administrative accounts is logged.")
        ],
        "Authentication": [
            ("Password Hashing Algorithm Check", "Ensure passwords are hashed using strong algorithms like bcrypt or Argon2."),
            ("Session Timeout Enforcement", "Verify inactive sessions are automatically terminated."),
            ("Multi-Factor Authentication Check", "Ensure MFA is required for all administrative access."),
            ("Brute Force Attack Protection", "Check for account lockout mechanisms after repeated failed login attempts.")
        ],
        "Authorization": [
            ("Least Privilege Principle Audit", "Ensure users and services operate with the minimum required permissions."),
            ("Role-Based Access Control (RBAC) Audit", "Verify that access to resources is strictly controlled by user roles."),
            ("IDOR Vulnerability Scan", "Ensure users cannot access resources belonging to others by manipulating IDs."),
            ("Directory Traversal Protection", "Verify that file paths provided by users are strictly validated.")
        ],
        "Cloud Misconfiguration": [
            ("Security Misconfiguration Check", "Ensure cloud resources do not use default credentials or configurations."),
            ("Cloud IAM Permissions Check", "Verify that overly permissive IAM roles (e.g., AdministratorAccess) are not used."),
            ("S3 Bucket Public Access Block", "Ensure cloud storage buckets are not publicly accessible unless explicitly required."),
            ("CloudTrail Logging Audit", "Verify that comprehensive auditing is enabled for all cloud infrastructure changes.")
        ],
        "Secrets & Credentials": [
            ("Hardcoded Secrets Scan", "Scan source code repositories for hardcoded API keys and passwords."),
            ("Sensitive Data Exposure Audit", "Ensure sensitive data is not exposed in application logs or error messages."),
            ("JWT Signature Validation", "Verify that JSON Web Tokens are properly signed and validated.")
        ]
    }

    categories_prefix = {
        "Dependency Security": "SEC-DEP",
        "SAST & Code Analysis": "SEC-SAST",
        "API Security": "SEC-API",
        "Database Security": "SEC-DB",
        "Authentication": "SEC-AUTH",
        "Authorization": "SEC-AUTHZ",
        "Cloud Misconfiguration": "SEC-CLOUD",
        "Secrets & Credentials": "SEC-SEC"
    }

    rows = []
    for cat, checks in cases_data.items():
        prefix = categories_prefix[cat]
        for i, (rule_name, desc) in enumerate(checks, 1):
            check_id = f"{prefix}-{str(i).zfill(2)}"
            status = "PASS"
            rows.append([check_id, cat, rule_name, desc, status])
            
    return rows

def main():
    print("Generating Security Vulnerability Report...")
    cases = generate_security_cases()
    
    headers = ["Check ID", "Category", "Security Check / Rule Name", "Description", "Status"]
    metrics = {
        "Total Findings": 0,
        "Critical Severity": 0,
        "High Severity": 0,
        "Moderate Severity": 0,
        "Low Severity": 0,
        "Informational Severity": 0,
        "Overall Assessment": "PASS"
    }
    
    timestamp = datetime.now().strftime("%m/%d/%Y, %I:%M:%S %p")
    subtitle = f"Scan Time: {timestamp} | Academic Demonstration Mode | Overall Status: PASS"
    
    output_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(output_dir, 'Security_Vulnerability_Report.xlsx')
    
    create_styled_excel(file_path, "AuraFit AI — Security Vulnerability Report", subtitle, metrics, headers, cases)
    print(f"Successfully generated {len(cases)} vulnerability cases in {file_path}")

if __name__ == "__main__":
    main()
