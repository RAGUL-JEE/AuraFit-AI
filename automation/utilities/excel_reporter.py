import os
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_excel_report(results_list):
    """
    Generates a highly formatted Excel report matching the requested MedMonitor aesthetic.
    """
    report_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reports')
    os.makedirs(report_dir, exist_ok=True)
    file_path = os.path.join(report_dir, 'Execution_Report.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Selenium Test Report"
    
    df = pd.DataFrame(results_list)
    
    if df.empty:
        wb.save(file_path)
        return
        
    total = len(df)
    passed = len(df[df['Status'] == 'Passed'])
    failed = len(df[df['Status'] == 'Failed'])
    skipped = len(df[df['Status'] == 'Skipped'])
    overall_status = "PASS" if failed == 0 else "FAIL"
    
    # --- Define Colors and Fonts ---
    title_fill = PatternFill(start_color="34D399", end_color="34D399", fill_type="solid") # Bright Green
    subtitle_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Very Dark Slate
    dark_bg_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    table_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Medium Slate
    
    pass_bg = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light Green BG
    pass_font = Font(color="047857", bold=True) # Dark Green Text
    
    fail_bg = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red BG
    fail_font = Font(color="B91C1C", bold=True) # Dark Red Text
    
    skip_bg = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Yellow BG
    skip_font = Font(color="B45309", bold=True) # Dark Yellow Text
    
    white_font = Font(color="FFFFFF", bold=True)
    white_font_normal = Font(color="FFFFFF")
    title_font = Font(size=16, bold=True, color="000000")
    subtitle_font = Font(size=10, italic=True, color="94A3B8")
    
    thin_border = Border(left=Side(style='thin', color="CBD5E1"), 
                         right=Side(style='thin', color="CBD5E1"), 
                         top=Side(style='thin', color="CBD5E1"), 
                         bottom=Side(style='thin', color="CBD5E1"))
                         
    no_border = Border()

    # --- Setup Layout ---
    # Col A: Vertical Summary Label (Width 5)
    # Col B: Metric Label (Width 25)
    # Col C: Metric Value (Width 15)
    # Col D: Description / Error Msg (Width 50)
    # Col E: Status (Width 15)
    # Col F, G: Extra details for the table below (Browser, Duration)
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15

    # 1. Title Banner
    ws.merge_cells('A1:G2')
    title_cell = ws['A1']
    title_cell.value = "AuraFit AI — Selenium Execution Report"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 2. Subtitle Banner
    ws.merge_cells('A3:G3')
    sub_cell = ws['A3']
    sub_cell.value = f"Execution Time: {datetime.datetime.now().strftime('%m/%d/%Y, %I:%M:%S %p')} | Environment: QA-Selenium | Overall Status: {overall_status}"
    sub_cell.font = subtitle_font
    sub_cell.fill = subtitle_fill
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Spacer
    ws.append([])
    
    # 3. Summary Header
    ws.merge_cells('A5:G5')
    sum_hdr = ws['A5']
    sum_hdr.value = "Execution Summary Metrics"
    sum_hdr.font = white_font
    sum_hdr.fill = dark_bg_fill
    sum_hdr.alignment = Alignment(horizontal='center', vertical='center')
    
    # 4. Summary Body
    summary_data = [
        ("Total Tests Executed", str(total), "Total number of UI automation scenarios run."),
        ("Passed Tests", str(passed), "Tests that completed successfully without errors."),
        ("Failed Tests", str(failed), "Tests that failed assertions or crashed."),
        ("Skipped Tests", str(skipped), "Tests intentionally skipped."),
        ("Pass Rate", f"{round((passed/total)*100, 2) if total else 0}%", "Percentage of successful tests."),
        ("Overall Assessment", overall_status, "Final status of the test suite.")
    ]
    
    start_row = 6
    for i, (label, val, desc) in enumerate(summary_data):
        row_idx = start_row + i
        ws.cell(row=row_idx, column=2, value=label).font = white_font_normal
        ws.cell(row=row_idx, column=2).fill = dark_bg_fill
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='right', vertical='center')
        
        ws.cell(row=row_idx, column=3, value=val).font = white_font
        ws.cell(row=row_idx, column=3).fill = dark_bg_fill
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=7)
        desc_cell = ws.cell(row=row_idx, column=4, value=desc)
        desc_cell.font = Font(color="94A3B8", italic=True)
        desc_cell.fill = dark_bg_fill
        desc_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Fill remaining merged cells with dark bg
        for col in range(5, 8):
            ws.cell(row=row_idx, column=col).fill = dark_bg_fill

    # Vertical SUMMARY text
    ws.merge_cells(f'A6:A{start_row + len(summary_data) - 1}')
    vert_cell = ws['A6']
    vert_cell.value = "SUMMARY"
    vert_cell.font = Font(color="34D399", bold=True, size=12) # Green text
    vert_cell.fill = dark_bg_fill
    vert_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

    # Spacer
    ws.append([])
    
    # 5. Detailed Test Cases Header
    detail_hdr_row = ws.max_row + 1
    ws.merge_cells(start_row=detail_hdr_row, start_column=1, end_row=detail_hdr_row, end_column=7)
    det_hdr = ws.cell(row=detail_hdr_row, column=1, value="DETAILED SELENIUM TEST CASES")
    det_hdr.font = white_font
    det_hdr.fill = dark_bg_fill
    det_hdr.alignment = Alignment(horizontal='center', vertical='center')
    
    # 6. Table Columns
    tbl_hdr_row = detail_hdr_row + 1
    headers = ["Test ID", "Module", "Feature/Test Name", "Error Message", "Status", "Browser", "Duration (s)"]
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=tbl_hdr_row, column=col_idx, value=hdr)
        cell.font = white_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    # 7. Table Data
    current_row = tbl_hdr_row + 1
    for item in results_list:
        # Standardize keys depending on where the data came from
        test_id = item.get("Test Case ID") or item.get("Test ID", "")
        module = item.get("Module", "")
        feature = item.get("Feature", "")
        error = item.get("Error Message", "")
        status = item.get("Status", "")
        browser = item.get("Browser", "")
        duration = item.get("Duration (s)") or item.get("Duration", "")
        
        row_data = [test_id, module, feature, error, status, browser, duration]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx != 4 else 'left', vertical='center')
            
            # Format Test ID
            if col_idx == 1:
                cell.font = Font(color="2563EB") # Blue link style
            
            # Format Status
            if col_idx == 5:
                if status == "Passed":
                    cell.fill = pass_bg
                    cell.font = pass_font
                elif status == "Failed":
                    cell.fill = fail_bg
                    cell.font = fail_font
                elif status == "Skipped":
                    cell.fill = skip_bg
                    cell.font = skip_font
                    
            # Very light grey alternating rows for other columns
            if col_idx != 5:
                if current_row % 2 == 0:
                    cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                    
        current_row += 1
        
    wb.save(file_path)
    print(f"Excel report generated successfully at {file_path}")

if __name__ == "__main__":
    # Mock data to test the layout
    sample = []
    for i in range(1, 15):
        status = "Passed"
        error = ""
        if i == 5 or i == 12:
            status = "Failed"
            error = "NoSuchElementException: unable to locate element"
        elif i == 8:
            status = "Skipped"
            error = "Test skipped due to missing prerequisite"
            
        sample.append({
            "Test ID": f"SEL-TC-{str(i).zfill(3)}",
            "Module": "Authentication" if i < 7 else "Dashboard",
            "Feature": "test_invalid_login" if i < 7 else "test_workout_detection",
            "Error Message": error,
            "Status": status,
            "Browser": "Chrome Headless",
            "Duration (s)": f"{5.2 + (i*0.1):.2f}"
        })
    generate_excel_report(sample)
